from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from config.defaults import DEFAULT_SETTINGS
from engine_pro import audit_dataframe
from utils.columns import auto_mapping, detected_groups
from utils.export_pro import export_workbook

df = pd.DataFrame({
    "Organisation unit name": ["مركز أ", "مركز أ", "مركز ب"],
    "رقم تعريف المريض": ["1", "", "3"],
    "نوع الإقامة": ["مقيم", "خطأ", "نازح"],
    "نوع الزيارة": ["جديدة", "", "مراجعة"],
    "تاريخ الميلاد": ["2010-01-01"] * 3,
    "Event date": ["2026-01-01"] * 3,
    "تشاخيص1": ["التهاب", "", ""],
})
ctx = {"filename":"sample.xlsx","clinic":"أطفال","mapping":auto_mapping(df.columns),"groups":detected_groups(df.columns),"settings":dict(DEFAULT_SETTINGS)}
errors, skipped = audit_dataframe(df, ctx)
assert "اكتمال تشخيص الأطفال" in set(errors["اسم قاعدة التدقيق"])
summary = {"نوع العيادة":"أطفال","إجمالي السجلات":3,"السجلات المتأثرة":2,"السجلات السليمة":1,"إجمالي الأخطاء والملاحظات":len(errors),"القواعد المتجاوزة":len(skipped),"نسبة السجلات المتأثرة":66.7}
blob = export_workbook(errors, skipped, summary)
wb = load_workbook(BytesIO(blob), read_only=False)
assert {"الملخص","إحصاء حسب القاعدة","إحصاء حسب المركز","جميع الأخطاء","القواعد المتجاوزة"}.issubset(wb.sheetnames)
assert len(wb["إحصاء حسب القاعدة"]._charts) == 1
app = AppTest.from_file("app_pro.py", default_timeout=15).run()
assert not app.exception
assert app.file_uploader
print(f"OK pro: errors={len(errors)}, sheets={len(wb.sheetnames)}, uploader={len(app.file_uploader)}")
