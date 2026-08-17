from io import BytesIO
import zipfile
import pandas as pd
from openpyxl import load_workbook
from utils.export_compat import export_workbook

def test_compatible_export_has_no_drawing_parts():
    errors=pd.DataFrame([{"اسم قاعدة التدقيق":"تشخيص NCD مفقود","درجة الحالة":"خطأ","Organisation unit name":"مركز","سبب الخطأ":"فارغ"}])
    blob=export_workbook(errors,pd.DataFrame(),{"نوع العيادة":"داخلية / NCD","إجمالي السجلات":1})
    archive=zipfile.ZipFile(BytesIO(blob))
    assert not any(name.startswith("xl/drawings/") for name in archive.namelist())
    wb=load_workbook(BytesIO(blob),read_only=False)
    assert "جميع الأخطاء" in wb.sheetnames and len(wb["إحصاء حسب القاعدة"]._charts)==0
