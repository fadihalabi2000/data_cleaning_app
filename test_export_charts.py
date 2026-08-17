from io import BytesIO
import zipfile

import pandas as pd
from openpyxl import load_workbook

from utils.export_charts import export_workbook


def test_chart_report_is_valid_and_contains_two_native_charts():
    errors = pd.DataFrame([
        {"اسم قاعدة التدقيق": "غياب التشخيص", "درجة الحالة": "خطأ", "Organisation unit name": "مركز أ"},
        {"اسم قاعدة التدقيق": "نوع الإقامة", "درجة الحالة": "خطأ", "Organisation unit name": "مركز ب"},
        {"اسم قاعدة التدقيق": "غياب التشخيص", "درجة الحالة": "خطأ", "Organisation unit name": "مركز أ"},
    ])
    blob = export_workbook(errors, pd.DataFrame(), {"نوع العيادة": "أطفال", "إجمالي السجلات": 10})
    with zipfile.ZipFile(BytesIO(blob)) as archive:
        assert archive.testzip() is None
        assert len([name for name in archive.namelist() if name.startswith("xl/charts/chart")]) == 2
    workbook = load_workbook(BytesIO(blob), read_only=False)
    assert "لوحة الإحصائيات" in workbook.sheetnames
    assert len(workbook["لوحة الإحصائيات"]._charts) == 2
