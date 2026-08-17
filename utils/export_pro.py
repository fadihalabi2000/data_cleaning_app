from io import BytesIO
import re
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY, TEAL, PALE, RED, AMBER = "123047", "087E8B", "E8F3F5", "D64550", "F5B700"

def safe_sheet_name(name, used):
    base = re.sub(r"[\\/*?:\[\]]", "-", str(name))[:31] or "Sheet"
    candidate, i = base, 2
    while candidate in used:
        suffix = f"-{i}"; candidate = base[:31-len(suffix)] + suffix; i += 1
    used.add(candidate); return candidate

def style_table(ws, reason=True):
    ws.sheet_view.rightToLeft = True; ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False; ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY); cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, header in enumerate(ws[1], 1):
        width = min(max(13, max((len(str(ws.cell(r, i).value or "")) for r in range(1, min(ws.max_row, 180)+1)), default=10)+2), 48)
        ws.column_dimensions[get_column_letter(i)].width = width
        if reason and header.value == "سبب الخطأ":
            for r in range(2, ws.max_row+1): ws.cell(r, i).fill = PatternFill("solid", fgColor="FFF4CC")
    for row in ws.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=False)

def export_workbook(errors, skipped, summary):
    output, used = BytesIO(), set()
    by_rule = errors.groupby(["اسم قاعدة التدقيق", "درجة الحالة"]).size().reset_index(name="العدد") if not errors.empty else pd.DataFrame(columns=["اسم قاعدة التدقيق","درجة الحالة","العدد"])
    by_center = errors.groupby("Organisation unit name").size().reset_index(name="العدد").sort_values("العدد", ascending=False) if not errors.empty else pd.DataFrame(columns=["Organisation unit name","العدد"])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary.items(), columns=["المؤشر", "القيمة"]).to_excel(writer, sheet_name="الملخص", index=False); used.add("الملخص")
        by_rule.to_excel(writer, sheet_name="إحصاء حسب القاعدة", index=False); used.add("إحصاء حسب القاعدة")
        by_center.to_excel(writer, sheet_name="إحصاء حسب المركز", index=False); used.add("إحصاء حسب المركز")
        (errors if not errors.empty else pd.DataFrame(columns=["لا توجد أخطاء"])).to_excel(writer, sheet_name="جميع الأخطاء", index=False); used.add("جميع الأخطاء")
        (skipped if not skipped.empty else pd.DataFrame(columns=["لا توجد قواعد متجاوزة"])).to_excel(writer, sheet_name="القواعد المتجاوزة", index=False); used.add("القواعد المتجاوزة")
        if not errors.empty:
            for rule, part in errors.groupby("اسم قاعدة التدقيق", sort=False): part.to_excel(writer, sheet_name=safe_sheet_name(rule, used), index=False)
        for ws in writer.book.worksheets: style_table(ws)
        summary_ws = writer.book["الملخص"]
        summary_ws.column_dimensions["A"].width = 38; summary_ws.column_dimensions["B"].width = 20
        if not by_rule.empty:
            chart_ws = writer.book["إحصاء حسب القاعدة"]
            chart = BarChart(); chart.type = "bar"; chart.style = 10; chart.title = "الأخطاء حسب القاعدة"; chart.height = 8; chart.width = 16
            chart.add_data(Reference(chart_ws, min_col=3, min_row=1, max_row=chart_ws.max_row), titles_from_data=True)
            chart.set_categories(Reference(chart_ws, min_col=1, min_row=2, max_row=chart_ws.max_row)); chart_ws.add_chart(chart, "E2")
    output.seek(0); return output.getvalue()

def export_single_rule(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="الأخطاء", index=False); style_table(writer.book["الأخطاء"])
    output.seek(0); return output.getvalue()
