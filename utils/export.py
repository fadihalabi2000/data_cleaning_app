from io import BytesIO
import re
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def safe_sheet_name(name, used):
    base=re.sub(r"[\\/*?:\[\]]","-",str(name))[:31] or "Sheet"
    candidate=base; i=2
    while candidate in used:
        suffix=f"-{i}"; candidate=base[:31-len(suffix)]+suffix; i+=1
    used.add(candidate); return candidate

def _style(ws):
    ws.sheet_view.rightToLeft=True; ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    fill=PatternFill("solid",fgColor="176B87")
    for cell in ws[1]:
        cell.fill=fill; cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    reason_col=None
    for i,c in enumerate(ws[1],1):
        if c.value=="سبب الخطأ": reason_col=i
        width=min(max(12,max((len(str(ws.cell(r,i).value or "")) for r in range(1,min(ws.max_row,150)+1)),default=12)+2),45)
        ws.column_dimensions[get_column_letter(i)].width=width
    if reason_col:
        for r in range(2,ws.max_row+1): ws.cell(r,reason_col).fill=PatternFill("solid",fgColor="FFF2CC")

def export_workbook(errors, skipped, summary):
    output=BytesIO(); used=set()
    with pd.ExcelWriter(output,engine="openpyxl") as writer:
        pd.DataFrame(summary.items(),columns=["المؤشر","القيمة"]).to_excel(writer,sheet_name="Summary",index=False); used.add("Summary")
        (errors if not errors.empty else pd.DataFrame(columns=["لا توجد أخطاء"])).to_excel(writer,sheet_name="All Errors",index=False); used.add("All Errors")
        (skipped if not skipped.empty else pd.DataFrame(columns=["لا توجد قواعد متجاوزة"])).to_excel(writer,sheet_name="Skipped Rules",index=False); used.add("Skipped Rules")
        if not errors.empty:
            for rule,part in errors.groupby("اسم قاعدة التدقيق",sort=False): part.to_excel(writer,sheet_name=safe_sheet_name(rule,used),index=False)
        for ws in writer.book.worksheets: _style(ws)
    output.seek(0); return output.getvalue()

def export_single_rule(df):
    output=BytesIO()
    with pd.ExcelWriter(output,engine="openpyxl") as writer:
        df.to_excel(writer,sheet_name="Errors",index=False); _style(writer.book["Errors"])
    output.seek(0); return output.getvalue()
