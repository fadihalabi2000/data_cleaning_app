from io import BytesIO

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill

from utils.export_pro import safe_sheet_name, style_table

ORANGE = "F36F3A"
BLACK = "000000"


def add_dashboard(writer, summary, by_rule, by_center):
    ws = writer.book.create_sheet("لوحة الإحصائيات", 1)
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F2")
    title = ws["A1"]
    title.value = "لوحة إحصائيات تدقيق البيانات"
    title.font = Font(size=20, bold=True, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=BLACK)
    title.alignment = Alignment(horizontal="center", vertical="center")

    for i, (label, value) in enumerate(list(summary.items())[:6]):
        col = (i % 3) * 2 + 1
        row = 4 + (i // 3) * 3
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col, label).font = Font(bold=True, color="666666")
        ws.cell(row + 1, col, value).font = Font(size=18, bold=True, color=ORANGE)
        ws.cell(row, col).alignment = ws.cell(row + 1, col).alignment = Alignment(horizontal="center")

    start = 11
    headers = [(1, "اسم قاعدة التدقيق"), (2, "العدد"), (4, "المركز"), (5, "العدد")]
    for col, value in headers:
        cell = ws.cell(start, col, value)
        cell.fill = PatternFill("solid", fgColor=BLACK)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    rules = by_rule.head(20)
    centers = by_center.head(15)
    for row, record in enumerate(rules.itertuples(index=False), start + 1):
        ws.cell(row, 1, record[0]); ws.cell(row, 2, int(record[1]))
    for row, record in enumerate(centers.itertuples(index=False), start + 1):
        ws.cell(row, 4, record[0] or "غير محدد"); ws.cell(row, 5, int(record[1]))

    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 4; ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 12

    if not rules.empty:
        chart = BarChart(); chart.type = "bar"; chart.style = 10
        chart.title = "الأخطاء حسب القاعدة"; chart.x_axis.title = "عدد الأخطاء"
        chart.height = 9; chart.width = 17; chart.legend = None
        chart.add_data(Reference(ws, min_col=2, min_row=start, max_row=start + len(rules)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=start + 1, max_row=start + len(rules)))
        chart.dLbls = DataLabelList(); chart.dLbls.showVal = True
        chart.series[0].graphicalProperties.solidFill = ORANGE
        chart.series[0].graphicalProperties.line.solidFill = ORANGE
        ws.add_chart(chart, "G3")

    if not centers.empty:
        chart = BarChart(); chart.type = "col"; chart.style = 10
        chart.title = "الأخطاء حسب المركز"; chart.y_axis.title = "عدد الأخطاء"
        chart.height = 9; chart.width = 17; chart.legend = None
        chart.add_data(Reference(ws, min_col=5, min_row=start, max_row=start + len(centers)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=4, min_row=start + 1, max_row=start + len(centers)))
        chart.dLbls = DataLabelList(); chart.dLbls.showVal = True
        chart.series[0].graphicalProperties.solidFill = ORANGE
        chart.series[0].graphicalProperties.line.solidFill = ORANGE
        ws.add_chart(chart, "G21")


def export_workbook(errors, skipped, summary):
    output, used = BytesIO(), set()
    if errors.empty:
        by_rule = pd.DataFrame(columns=["اسم قاعدة التدقيق", "العدد"])
        by_rule_severity = pd.DataFrame(columns=["اسم قاعدة التدقيق", "درجة الحالة", "العدد"])
        by_center = pd.DataFrame(columns=["Organisation unit name", "العدد"])
    else:
        by_rule = errors.groupby("اسم قاعدة التدقيق").size().reset_index(name="العدد").sort_values("العدد", ascending=False)
        by_rule_severity = errors.groupby(["اسم قاعدة التدقيق", "درجة الحالة"]).size().reset_index(name="العدد")
        by_center = errors.groupby("Organisation unit name").size().reset_index(name="العدد").sort_values("العدد", ascending=False)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary.items(), columns=["المؤشر", "القيمة"]).to_excel(writer, sheet_name="الملخص", index=False); used.add("الملخص")
        by_rule_severity.to_excel(writer, sheet_name="إحصاء حسب القاعدة", index=False); used.add("إحصاء حسب القاعدة")
        by_center.to_excel(writer, sheet_name="إحصاء حسب المركز", index=False); used.add("إحصاء حسب المركز")
        (errors if not errors.empty else pd.DataFrame(columns=["لا توجد أخطاء"])).to_excel(writer, sheet_name="جميع الأخطاء", index=False); used.add("جميع الأخطاء")
        (skipped if not skipped.empty else pd.DataFrame(columns=["لا توجد قواعد متجاوزة"])).to_excel(writer, sheet_name="القواعد المتجاوزة", index=False); used.add("القواعد المتجاوزة")
        if not errors.empty:
            for rule, part in errors.groupby("اسم قاعدة التدقيق", sort=False):
                part.to_excel(writer, sheet_name=safe_sheet_name(rule, used), index=False)
        for ws in writer.book.worksheets:
            style_table(ws)
        add_dashboard(writer, summary, by_rule, by_center)
        writer.book.calculation.fullCalcOnLoad = True
        writer.book.calculation.forceFullCalc = True
    output.seek(0)
    return output.getvalue()
